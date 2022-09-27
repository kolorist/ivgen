import openpyxl
import xml.dom.minidom

def entry_point():
    print("invoice_generator v0")

    doc = xml.dom.minidom.parse("invoices/849804_C22TGM_20092022/849804.xml")
    hdon = doc.getElementsByTagName("HDon")[0]
    dlhdon = hdon.getElementsByTagName("DLHDon")[0]

    ttchung = dlhdon.getElementsByTagName("TTChung")[0]
    shdon = ttchung.getElementsByTagName("SHDon")[0].firstChild.data
    nlap = ttchung.getElementsByTagName("NLap")[0].firstChild.data

    ndhdon = dlhdon.getElementsByTagName("NDHDon")[0]
    dshhdvu = ndhdon.getElementsByTagName("DSHHDVu")[0]
    hhdvuArr = dshhdvu.getElementsByTagName("HHDVu")

    ttoan = ndhdon.getElementsByTagName("TToan")[0]
    tienTruocThue = ttoan.getElementsByTagName("TgTCThue")[0].firstChild.data
    tienThue = ttoan.getElementsByTagName("TgTThue")[0].firstChild.data
    tienTong = ttoan.getElementsByTagName("TgTTTBSo")[0].firstChild.data

    mhhdvuList = []
    mhhdvuListStr = ""
    for hhdvu in hhdvuArr:
        mhhdvu = hhdvu.getElementsByTagName("MHHDVu")[0]
        if mhhdvu.firstChild.data not in mhhdvuList:
            if len(mhhdvuList) > 0:
                mhhdvuListStr += ",\n"
            mhhdvuList.append(mhhdvu.firstChild.data)
            mhhdvuListStr += mhhdvu.firstChild.data

    wb = openpyxl.load_workbook("template.xlsx")
    print(wb.sheetnames)

    ws = wb['Sheet1']
    ws['C8'] = 1
    ws['D8'] = nlap
    ws['J8'] = shdon
    ws['E8'] = mhhdvuListStr
    ws['F8'] = "Huong"
    ws['G8'].number_format = '#,##0'
    ws['G8'] = float(tienTruocThue)
    ws['H8'] = float(tienThue)
    ws['H8'].number_format = '#,##0'
    ws['K8'] = float(tienTong)
    ws['K8'].number_format = '#,##0'

    # c0 = ws['A3']
    # print(c0.value)
    # ws['A5'] = "1\n2\n3"
    # ws['A5'].alignment = openpyxl.styles.Alignment(wrapText=True)

    # ws.insert_rows(5, 3)

    wb.save("output.xlsx")


if __name__ == "__main__":
    entry_point()
